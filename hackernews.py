import requests
import re
import feedparser
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ================= CONFIG =================

RSS_URL = "https://feeds.feedburner.com/TheHackersNews"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

CVE_PATTERN = r"CVE[\s\-–—]*\d{4}[\s\-–—]*\d{4,7}"

FILE = "data.xlsx"

# ================= HELPERS =================

def clean_cve(cve):

    return re.sub(
        r"[\s–—\-]+",
        "-",
        cve.upper()
    ).strip("-")

# =========================================================
# EXTRACT CVEs
# =========================================================

def extract_cves(url):

    try:

        res = requests.get(
            url,
            headers=HEADERS,
            timeout=10
        )

    except:

        return set()

    soup = BeautifulSoup(
        res.text,
        "html.parser"
    )

    body = (
        soup.find("div", class_="articlebody")
        or
        soup.find("div", class_="post-body")
    )

    text = (
        body.get_text()
        if body
        else soup.get_text()
    )

    raw = re.findall(
        CVE_PATTERN,
        text,
        re.IGNORECASE
    )

    return {
        clean_cve(c)
        for c in raw
    }

# =========================================================
# LOAD EXISTING ROWS
# =========================================================

def load_existing_rows():

    if not os.path.exists(FILE):
        return []

    wb = load_workbook(FILE)

    ws = wb.active

    rows = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:

            rows.append(row)

    return rows

# =========================================================
# DUPLICATE LOGIC
# SAME CVE + SAME DATE + SAME LINK = SKIP
# =========================================================

def load_existing_keys():

    rows = load_existing_rows()

    return {
        (
            str(r[0]),  # CVE
            str(r[1]),  # DATE
            str(r[2])   # LINK
        )
        for r in rows
    }

# =========================================================
# SAVE EXCEL
# =========================================================

def save_all(rows):

    wb = Workbook()

    ws = wb.active

    ws.append([
        "CVE",
        "DATE",
        "LINK"
    ])

    # =====================================================
    # SORT OLDEST → NEWEST
    # =====================================================

    rows_sorted = sorted(
        rows,
        key=lambda x: datetime.strptime(
            x[1],
            "%Y-%m-%d"
        )
    )

    for r in rows_sorted:

        ws.append(r)

    wb.save(FILE)

# =========================================================
# MAIN
# =========================================================

def main():

    print(
        "\nRunning HackerNews CVE fetch...\n"
    )

    feed = feedparser.parse(RSS_URL)

    existing_rows = load_existing_rows()

    existing_keys = load_existing_keys()

    new_rows = []

    for entry in feed.entries:

        if not hasattr(
            entry,
            "published_parsed"
        ):
            continue

        pub = datetime(
            *entry.published_parsed[:6]
        )

        date_str = pub.strftime(
            "%Y-%m-%d"
        )

        link = entry.link

        print("\n" + "=" * 60)
        print("ARTICLE :", entry.title)
        print("DATE    :", date_str)

        cves = extract_cves(link)

        if cves:

            print("CVEs    :", cves)

        else:

            print("No CVEs Found")

        # =================================================
        # ONE CVE = ONE ROW
        # =================================================

        for c in cves:

            key = (
                c,
                date_str,
                link
            )

            # =============================================
            # SKIP ONLY EXACT SAME OBSERVATION
            # =============================================

            if key not in existing_keys:

                new_rows.append(
                    (
                        c,
                        date_str,
                        link
                    )
                )

                existing_keys.add(key)

    # =====================================================
    # APPEND HISTORY
    # =====================================================

    all_rows = existing_rows + new_rows

    if new_rows:

        save_all(all_rows)

        print("\n" + "=" * 60)
        print(
            "NEW ROWS ADDED :",
            len(new_rows)
        )
        print(
            "TOTAL ROWS     :",
            len(all_rows)
        )
        print("=" * 60)

    else:

        print("\nNo new data")

# ================= RUN =================

if __name__ == "__main__":

    main()